"""Parser for XLSX question and answer sheets."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import TYPE_CHECKING, Any, Iterable, Iterator, Sequence, cast

try:  # pragma: no cover - optional runtime dependency
    from openpyxl import load_workbook as _load_workbook
except Exception:  # pragma: no cover - graceful fallback for typing
    _load_workbook = None  # type: ignore[assignment]

load_workbook: Any = _load_workbook

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet
else:  # pragma: no cover - runtime fallback
    Worksheet = Any  # type: ignore[assignment]

from ..models import ParsedDocument, ParsedSection

EXPECTED_COLUMNS = {"question", "answer"}


def _ensure_worksheet(candidate: Any) -> Worksheet | None:
    """Return a Worksheet when the object quacks like one, otherwise None."""
    if candidate is None:
        return None
    if hasattr(candidate, "iter_rows") and hasattr(candidate, "max_row"):
        return cast(Worksheet, candidate)
    return None


def _cell_to_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, (int, float)):
        text = f"{value}"
        return text[:-2] if text.endswith(".0") else text
    return str(value).strip()


def _coerce_int(value: str) -> int | None:
    text = value.strip()
    if not text:
        return None
    try:
        return int(text)
    except ValueError:
        try:
            return int(float(text))
        except ValueError:
            return None


def _cell_at(cells: Sequence[str], index: int) -> str:
    if 0 <= index < len(cells):
        return cells[index]
    return ""


@dataclass(slots=True)
class Row:
    cells: list[str]


def iter_rows_safe(ws: Worksheet | None) -> Iterator[Row]:
    sheet = _ensure_worksheet(ws)
    if sheet is None:
        return
    try:
        values_iter = cast(Iterable[Sequence[Any]], sheet.iter_rows(values_only=True))
        for raw in values_iter:
            yield Row([_cell_to_str(value) for value in raw])
    except Exception:  # pragma: no cover - defensive fallback
        cells_iter = cast(Iterable[Sequence[Any]], sheet.iter_rows())
        for raw in cells_iter:
            yield Row([_cell_to_str(getattr(cell, "value", None)) for cell in raw])


def _header_map(header_row: Sequence[str]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for idx, cell in enumerate(header_row):
        key = cell.strip().lower()
        if key:
            mapping[key] = idx
    return mapping


def parse_xlsx(path: Path) -> ParsedDocument:
    if load_workbook is None:  # pragma: no cover - import guard
        raise RuntimeError("openpyxl is required to parse XLSX documents")

    workbook = load_workbook(filename=str(path), data_only=True)

    candidate_sheets: list[Worksheet | None] = []
    active = _ensure_worksheet(getattr(workbook, "active", None))
    if active is not None:
        candidate_sheets.append(active)
    for name in getattr(workbook, "sheetnames", []):
        try:
            sheet = _ensure_worksheet(workbook[name])
        except Exception:  # pragma: no cover - defensive guard
            sheet = None
        if sheet is not None and sheet not in candidate_sheets:
            candidate_sheets.append(sheet)

    rows: list[Row] = []
    for sheet in candidate_sheets:
        rows = list(iter_rows_safe(sheet))
        if rows:
            break

    if not rows:
        return ParsedDocument(source_path=path, source_type="xlsx", sections=[])

    headers = rows[0].cells
    header_map = _header_map(headers)
    column_map = {name: header_map[name] for name in EXPECTED_COLUMNS if name in header_map}
    sections: list[ParsedSection] = []

    for row in rows[1:]:
        cells = row.cells
        if not any(cells):
            continue

        question_idx = column_map.get("question")
        answer_idx = column_map.get("answer")
        question = _cell_at(cells, question_idx).strip() if question_idx is not None else ""
        answer = _cell_at(cells, answer_idx).strip() if answer_idx is not None else ""

        if not question and not answer:
            continue

        extra: dict[str, str] = {}
        source_idx = header_map.get("source")
        if source_idx is not None:
            source = _cell_at(cells, source_idx).strip()
            if source:
                extra["source"] = source

        page = None
        page_idx = header_map.get("page")
        if page_idx is not None:
            page = _coerce_int(_cell_at(cells, page_idx))

        text = f"Q: {question}\nA: {answer}"
        breadcrumbs = ["Q&A"]
        sections.append(
            ParsedSection(text=text, page_number=page, extra=extra, breadcrumbs=breadcrumbs)
        )

    return ParsedDocument(source_path=path, source_type="xlsx", sections=sections)
